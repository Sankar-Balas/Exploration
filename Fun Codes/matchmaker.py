import openpyxl

SourceBook = ""
SourceSheet = 0
OrphanFlag = False

def main():
    global SourceBook
    global SourceSheet
    try:
        SourceBook   = openpyxl.load_workbook('D:\\config\\HPALM_Requirements.xlsx')
        SourceSheet  = SourceBook.active
        SourceRowVal = 2        
        TotalRows    = 16489
        
        for x in range(TotalRows):        
            R2RC2TestCase       = SourceSheet.cell(row = SourceRowVal,column = 7).value
            MapExcelSheet(R2RC2TestCase)
            SourceRowVal = SourceRowVal + 1
            print (SourceRowVal)

    except Exception as e:
        print(e.args())
        
    finally:
        print("Done. Saving")
        SourceBook.save('D:\\config\\HPALM_Requirements.xlsx')

def MapExcelSheet(R2RC2TestCase):
    try:
        DestRowVal = 2
        if (R2RC2TestCase is not None):
            for row in SourceSheet.rows:
                MY20ManualTestCase    = SourceSheet.cell(row = DestRowVal,column = 4).value
                MY20AutoTestCase    = SourceSheet.cell(row = DestRowVal,column = 5).value
                MY20OtherTestCase = SourceSheet.cell(row = DestRowVal,column = 6).value
                
                if (MY20ManualTestCase is  None):   MY20ManualTestCase = ""
                if (MY20AutoTestCase is  None):   MY20AutoTestCase = ""
                if (MY20OtherTestCase is  None):   MY20OtherTestCase = ""
                
                if ((R2RC2TestCase in MY20ManualTestCase) or (R2RC2TestCase in MY20AutoTestCase) or (R2RC2TestCase in MY20OtherTestCase)):
                    SourceSheet.cell(row=DestRowVal,column=3).value = "Covered"
                    
                DestRowVal = DestRowVal + 1


    except Exception as e:
        print(e.args())
main()                